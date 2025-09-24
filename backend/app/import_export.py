"""Utilities for importing and exporting RoB 2 workbooks."""

from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Tuple

import yaml
from openpyxl import Workbook, load_workbook

from . import models, rule_engine

ROOT_DIR = Path(__file__).resolve().parents[2]
MAP_PATH = ROOT_DIR / "mapeamento.xlsx.yaml"


def load_mapping() -> dict:
    with open(MAP_PATH, "r", encoding="utf-8") as file_obj:
        return yaml.safe_load(file_obj)


_TOKEN_PATTERN = re.compile(r"(?P<key>[A-Za-z_]+)(?:\[(?P<index>[^\]]+)\])?")


def _split_path(path: str) -> List[Tuple[str, str | None]]:
    return [(match.group("key"), match.group("index")) for match in _TOKEN_PATTERN.finditer(path)]


def _ensure_domain(payload: Dict[int, Dict[str, Any]], domain_id: int) -> Dict[str, Any]:
    entry = payload.setdefault(domain_id, {})
    entry.setdefault("respostas", {})
    entry.setdefault("comentarios", None)
    entry.setdefault("observacoes_itens", {})
    entry.setdefault("direcao", "NA")
    return entry


def _assign_value(store: Dict[str, Any], path: str, value: Any, warnings: List[str]) -> None:
    tokens = _split_path(path)
    if not tokens:
        return

    root_key, root_index = tokens[0]

    if root_key == "Dominio":
        if root_index is None:
            warnings.append(f"Caminho sem índice de domínio: {path}")
            return
        domain_id = int(root_index)
        domain_entry = _ensure_domain(store.setdefault("dominios", {}), domain_id)
        _assign_nested(domain_entry, tokens[1:], value)
        return

    if root_key == "Resultado":
        target = store.setdefault("resultado", {})
        _assign_nested(target, tokens[1:], value)
        return

    if root_key == "AvaliacaoRob2":
        meta = store.setdefault("avaliacao_meta", {})
        if root_index:
            meta = meta.setdefault(root_key, {})
        _assign_nested(meta, tokens[1:], value)
        return

    if root_key == "Resumo":
        resumo = store.setdefault("resumo", {})
        remaining = path.split(".", 1)
        resumo_key = remaining[1] if len(remaining) == 2 else root_key
        resumo[resumo_key] = value
        return

    _assign_nested(store.setdefault(root_key, {}), tokens[1:], value, root_index)


def _assign_nested(current: Dict[str, Any], tokens: List[Tuple[str, str | None]], value: Any, initial_index: str | None = None) -> None:
    if initial_index is not None:
        current = current.setdefault(initial_index, {})

    for idx, (key, index) in enumerate(tokens):
        is_last = idx == len(tokens) - 1
        if is_last:
            if index is not None:
                target_dict = current.setdefault(key, {})
                target_dict[index] = value
            else:
                current[key] = value
        else:
            next_container = current.setdefault(key, {})
            if index is not None:
                next_container = next_container.setdefault(index, {})
            current = next_container


def import_workbook(file_path: Path) -> Tuple[Dict[str, Any], List[str]]:
    mapping = load_mapping()
    workbook = load_workbook(filename=str(file_path), data_only=True)
    warnings: List[str] = []

    payload: Dict[str, Any] = {"dominios": {}}

    for sheet_name, sheet_map in mapping.get("abas", {}).items():
        if sheet_name not in workbook.sheetnames:
            warnings.append(f"Aba '{sheet_name}' não encontrada no arquivo.")
            continue

        ws = workbook[sheet_name]
        if ws.max_row < 2:
            continue

        header_index = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
        sheet_row = None
        row_data: Dict[str, Any] = {}
        for row_idx in range(2, ws.max_row + 1):
            has_data = False
            current_row: Dict[str, Any] = {}
            for header in sheet_map.get("colunas", {}):
                col_idx = header_index.get(header)
                if not col_idx:
                    continue
                value = ws.cell(row=row_idx, column=col_idx).value
                if value not in (None, ""):
                    has_data = True
                current_row[header] = value
            if has_data:
                sheet_row = row_idx
                row_data = current_row
                break

        if not row_data:
            continue

        for header, path in sheet_map.get("colunas", {}).items():
            if header not in row_data:
                warnings.append(f"Coluna '{header}' ausente na aba '{sheet_name}'.")
                continue
            value = row_data.get(header)
            if value in (None, ""):
                continue
            _assign_value(payload, path, value, warnings)

    evaluation_payload = {
        "pre_consideracoes": payload.get("Pre_Consideracoes", {}).get("Observacoes")
        if isinstance(payload.get("Pre_Consideracoes"), dict)
        else payload.get("pre_consideracoes"),
        "dominios": [],
        "resultado": payload.get("resultado", {}),
    }

    for domain_id, data in sorted(payload.get("dominios", {}).items(), key=lambda item: item[0]):
        respostas = {k: v for k, v in (data.get("respostas") or {}).items() if v not in (None, "")}
        evaluation_payload["dominios"].append(
            {
                "tipo": domain_id,
                "respostas": respostas,
                "comentarios": data.get("comentarios"),
                "observacoes_itens": data.get("observacoes_itens") or {},
                "direcao": data.get("direcao") or "NA",
            }
        )

    return evaluation_payload, warnings


def export_workbook(result: models.Result, mapping: dict | None = None) -> Tuple[bytes, List[str]]:
    mapping = mapping or load_mapping()
    workbook = Workbook()
    warnings: List[str] = []

    # Remove default sheet created by openpyxl
    if workbook.sheetnames:
        workbook.remove(workbook.active)

    avaliacao = result.avaliacao
    dominios = {dom.tipo: dom for dom in (avaliacao.dominios if avaliacao else [])}

    for sheet_name, sheet_map in mapping.get("abas", {}).items():
        ws = workbook.create_sheet(title=sheet_name)
        headers = list(sheet_map.get("colunas", {}).keys())
        ws.append(headers)
        row_values = []
        for header in headers:
            path = sheet_map["colunas"].get(header)
            value = _resolve_value_for_path(result, dominios, avaliacao, path)
            if isinstance(value, list):
                value = ", ".join(str(item) for item in value if item is not None)
            row_values.append(value)
        ws.append(row_values)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), warnings


def _resolve_value_for_path(result: models.Result, dominios: Dict[int, models.Domain], avaliacao: models.Evaluation | None, path: str) -> Any:
    tokens = _split_path(path)
    if not tokens:
        return None
    root_key, root_index = tokens[0]

    if root_key == "Resultado":
        return _resolve_result_value(result, tokens[1:])

    if root_key == "Dominio":
        if root_index is None:
            return None
        domain = dominios.get(int(root_index))
        if not domain:
            return None
        return _resolve_domain_value(domain, tokens[1:])

    if root_key == "AvaliacaoRob2" and avaliacao:
        return _resolve_avaliacao_meta(avaliacao, tokens[1:])

    if root_key == "Resumo" and avaliacao:
        key = tokens[0][0]
        if len(tokens) >= 2:
            _, sub_key = tokens[1]
            target_key = sub_key or tokens[1][0]
        else:
            target_key = root_index or "valor"
        if target_key == "JulgamentoGlobal":
            return avaliacao.julgamento_global
        if target_key == "DirecaoGlobal":
            return avaliacao.direcao_global.value if avaliacao.direcao_global else None
        return None

    return None


def _resolve_result_value(result: models.Result, tokens: List[Tuple[str, str | None]]) -> Any:
    current: Any = result
    for key, index in tokens:
        if not current:
            return None
        if hasattr(current, key):
            current = getattr(current, key)
        elif isinstance(current, dict):
            if index:
                current = current.get(key, {}).get(index)
            else:
                current = current.get(key)
        else:
            current = None
    return current


def _resolve_domain_value(domain: models.Domain, tokens: List[Tuple[str, str | None]]) -> Any:
    current: Any = domain
    for key, index in tokens:
        if key == "respostas" and index is not None:
            return (domain.respostas or {}).get(index)
        if key == "comentarios":
            return domain.comentarios
        if key == "observacoes_itens" and index is not None:
            return (domain.observacoes_itens or {}).get(index)
        if key.lower().startswith("julgamento"):
            return domain.julgamento
        if key.lower().startswith("direcao"):
            return domain.direcao.value if domain.direcao else None
    return None


def _resolve_avaliacao_meta(avaliacao: models.Evaluation, tokens: List[Tuple[str, str | None]]) -> Any:
    if not tokens:
        return None
    key, index = tokens[0]
    if key.lower().startswith("julgamento") and index is not None:
        dom = next((d for d in avaliacao.dominios if str(d.tipo) == index), None)
        return dom.julgamento if dom else None
    if key == "julgamentoGlobal":
        return avaliacao.julgamento_global
    if key.lower().startswith("direcao"):
        return avaliacao.direcao_global.value if avaliacao.direcao_global else None
    return None


def persist_imported_evaluation(
    db,
    current_user: models.User,
    result: models.Result,
    evaluation_payload: Dict[str, Any]
) -> models.Evaluation:
    avaliacao = result.avaliacao
    if not avaliacao:
        avaliacao = models.Evaluation(resultado_id=result.id, criado_por_id=current_user.id)
        db.add(avaliacao)
        db.flush()

    resultado_data = evaluation_payload.get("resultado", {}) or {}
    if resultado_data.get("desfecho"):
        result.desfecho = resultado_data["desfecho"]
    if resultado_data.get("desenho"):
        result.desenho = resultado_data["desenho"]
    if resultado_data.get("resultadoNumerico"):
        result.resultado_numerico = resultado_data["resultadoNumerico"]
    if resultado_data.get("efeitoInteresse"):
        result.efeito_interesse = resultado_data["efeitoInteresse"]
    if resultado_data.get("medida_efeito"):
        result.medida_efeito = resultado_data["medida_efeito"]

    fontes_payload = resultado_data.get("fontes")
    extra_fontes = result.fontes.copy() if isinstance(result.fontes, dict) else {}
    for key in ("intervencaoExperimental", "comparador"):
        if resultado_data.get(key):
            extra_fontes[key] = resultado_data[key]
    if fontes_payload:
        if isinstance(fontes_payload, dict):
            extra_fontes.update(fontes_payload)
        else:
            extra_fontes["descricao"] = str(fontes_payload)
    result.fontes = extra_fontes or None

    avaliacao.pre_consideracoes = evaluation_payload.get("pre_consideracoes")

    avaliacao.dominios.clear()
    julgamentos: List[str] = []
    justifications: List[str] = []
    directions: List[models.DirectionType] = []

    for dominio in evaluation_payload.get("dominios", []):
        julgamento, justificativa = rule_engine.evaluate_domain(dominio["tipo"], dominio.get("respostas", {}))
        direcao_str = dominio.get("direcao") or "NA"
        try:
            direcao_enum = models.DirectionType(direcao_str)
        except ValueError:
            direcao_enum = models.DirectionType.NA

        novo_dominio = models.Domain(
            avaliacao_id=avaliacao.id,
            tipo=dominio["tipo"],
            respostas=dominio.get("respostas", {}),
            comentarios=dominio.get("comentarios"),
            observacoes_itens=dominio.get("observacoes_itens", {}),
            julgamento=julgamento,
            justificativa=justificativa,
            direcao=direcao_enum,
        )
        db.add(novo_dominio)
        julgamentos.append(julgamento)
        if justificativa:
            justifications.append(f"Domínio {dominio['tipo']}: {justificativa}")
        if direcao_enum != models.DirectionType.NA:
            directions.append(direcao_enum)

    avaliacao.julgamento_global = rule_engine.evaluate_global(julgamentos)
    avaliacao.direcao_global = directions[0] if directions else models.DirectionType.NA
    avaliacao.justificativa_global = "\n".join(justifications) if justifications else None

    db.commit()
    db.refresh(avaliacao)
    return avaliacao
