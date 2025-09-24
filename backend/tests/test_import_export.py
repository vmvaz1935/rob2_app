from backend.app import import_export


def test_Import_excel_deve_mapear_Q1_1_para_resposta_1_1():
    mapping = import_export.load_mapping()
    dominio1_cols = mapping['abas']['Dominio1']['colunas']
    assert dominio1_cols.get('Q1_1') == "Dominio[1].respostas[1.1]"


def test_Export_excel_deve_emitir_colunas_template():
    mapping = import_export.load_mapping()
    resumo = mapping['abas']['Resumo']['colunas']
    assert 'JulgamentoGlobal' in resumo
    assert resumo['JulgamentoGlobal'] == "AvaliacaoRob2.julgamentoGlobal"

from io import BytesIO
from openpyxl import Workbook, load_workbook

from backend.app import import_export, models


def criar_planilha_exemplo(caminho):
    wb = Workbook()
    ws_pre = wb.active
    ws_pre.title = "Pre_Consideracoes"
    ws_pre.append(["Desenho", "Experimental", "Comparador", "Desfecho", "ResultadoNumerico", "EfeitoInteresse", "Fontes"])
    ws_pre.append(["Paralelo", "Intervenção A", "Placebo", "Dor", "10%", "assignment", "Registro X"])

    ws_dom1 = wb.create_sheet("Dominio1")
    ws_dom1.append(["Q1_1", "Q1_2", "Q1_3", "Q1_4", "Comentarios", "Julgamento"])
    ws_dom1.append(["Y", "Y", "N", "NA", "Comentário 1", "Baixo"])

    ws_dom2 = wb.create_sheet("Dominio2")
    ws_dom2.append(["Q2_1", "Q2_2", "Q2_3", "Comentarios", "Julgamento"])
    ws_dom2.append(["N", "Y", "Y", "Comentário 2", "Baixo"])

    ws_dom3 = wb.create_sheet("Dominio3")
    ws_dom3.append(["Q3_1", "Q3_2", "Q3_3", "Comentarios", "Julgamento"])
    ws_dom3.append(["N", "N", "Y", "Comentário 3", "Baixo"])

    ws_dom4 = wb.create_sheet("Dominio4")
    ws_dom4.append(["Q4_1", "Q4_2", "Q4_3", "Comentarios", "Julgamento"])
    ws_dom4.append(["Y", "Y", "N", "Comentário 4", "Baixo"])

    ws_dom5 = wb.create_sheet("Dominio5")
    ws_dom5.append(["Q5_1", "Q5_2", "Comentarios", "Julgamento"])
    ws_dom5.append(["Y", "N", "Comentário 5", "Baixo"])

    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.append(["JulgamentoGlobal", "DirecaoGlobal"])
    ws_resumo.append(["Baixo", "NA"])

    wb.save(caminho)


def test_import_workbook_deve_retornar_dominios_e_resultado(tmp_path):
    arquivo = tmp_path / "avaliacao.xlsx"
    criar_planilha_exemplo(arquivo)

    payload, warnings = import_export.import_workbook(arquivo)

    assert warnings == []
    assert payload["resultado"]["desfecho"] == "Dor"
    dominio1 = next(item for item in payload["dominios"] if item["tipo"] == 1)
    assert dominio1["respostas"]["1.2"] == "Y"


def test_export_workbook_deve_gerar_planilha_compativel():
    estudo = models.Study(projeto_id=1, referencia="Estudo X", desenho="Paralelo")
    resultado = models.Result(estudo=estudo, desfecho="Dor", resultado_numerico="10%")
    avaliacao = models.Evaluation(resultado=resultado, pre_consideracoes="Observações", julgamento_global="Baixo", direcao_global=models.DirectionType.NA)
    dominio = models.Domain(tipo=1, respostas={"1.1": "Y", "1.2": "PY"}, comentarios="Comentário", observacoes_itens={"1.1": "Obs"}, julgamento="Baixo", justificativa="Randomização adequada")
    avaliacao.dominios = [dominio]
    resultado.avaliacao = avaliacao

    workbook_bytes, warnings = import_export.export_workbook(resultado)

    assert warnings == []
    wb = load_workbook(filename=BytesIO(workbook_bytes))
    assert "Pre_Consideracoes" in wb.sheetnames
    ws_pre = wb["Pre_Consideracoes"]
    headers = [cell.value for cell in ws_pre[1]]
    assert "Desfecho" in headers
    desfecho_idx = headers.index("Desfecho")
    assert ws_pre[2][desfecho_idx].value == "Dor"

